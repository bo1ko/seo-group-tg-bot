import random
import glob
import os
import asyncio

from openpyxl import load_workbook
from dotenv import load_dotenv


load_dotenv()

async def random_sleep():
    sleep_time = random.uniform(50, 60)
    await asyncio.sleep(sleep_time)


def load_excel_data():
    workbook = load_workbook(os.getenv('EXCEL'))
    sheet = workbook.active
    data = []

    for row in sheet.iter_rows(values_only=True):
        data.append(row)
        
    count = len(data) - 1

    return data, count
